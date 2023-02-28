from pprint import pprint
import os
import json
#import win32com.client as win32
from openpyxl import Workbook  
import time  

def create_excel(path, name):
    wb = Workbook()  
    sheet = wb.active
    sheet.title = 'Speaker Diarisation'

    #step 1.1- Read the json file
    json_data=json.loads(open(path).read())

    #insert headers
    header_labels=('speaker', 'text' , 'emotion', 'time (ms)', 'dbfs')

    for index, val in enumerate(header_labels):
        sheet.cell(row=1, column=index+1).value = val
    
    speakers=[]
    row_num = 2
    for i, data in enumerate(json_data):
        speakers.append(data['speaker'])
        for j, dbfs in enumerate(data['dbfs']):
            sheet.cell(row=row_num, column=1).value = data['speaker']
            sheet.cell(row=row_num, column=2).value = data['text']
            sheet.cell(row=row_num, column=3).value = data['emotion']
            sheet.cell(row=row_num, column=4).value = dbfs['time']
            sheet.cell(row=row_num, column=5).value = dbfs['value']
        
            row_num=row_num+1


    # Create speaker talktime distribution worksheet
    #print(speakers)

    s = ''
    sdecibel = []
    stime=[]
    avg = []
    for data in json_data:
        # Assign speaker value if empty
        if s == '':
            s = data['speaker']

        # If the speaker changes, add the total time and decibel to the collection.
        # Reset the value collection so that it starts calculating for the next speaker in the list.
        if s != data['speaker']:
            avg.append((s, sum(sdecibel)/len(sdecibel), stime[len(stime)-1]-stime[0]))
            sdecibel = []
            stime=[]
        
        for dbfs in data['dbfs']:
            sdecibel.append(dbfs['value'])
            stime.append(dbfs['time'])

        s = data['speaker']

    # Add avg of the time.
    avg.append((s, sum(sdecibel)/len(sdecibel), stime[len(stime)-1]-stime[0]))
    #print(avg)

    wb.create_sheet('Speaker Talktime Distribution') # Create a new sheet
    wb.active = wb['Speaker Talktime Distribution'] # Set it as active
    sheet1 = wb.active
    sheet1.cell(row=1, column=1).value = 'Speaker'
    sheet1.cell(row=1, column=2).value = 'Average time (sec)'
    sheet1.cell(row=1, column=3).value = 'Average dbfs'

    row_num2 = 2
    for val in avg:
        sheet1.cell(row=row_num2, column=1).value = val[0]
        sheet1.cell(row=row_num2, column=2).value = val[2]/1000
        sheet1.cell(row=row_num2, column=3).value = val[1]
        row_num2 = row_num2 +1
        
    # unique_speakers = unique(speakers)
   
    # st=[]
    # for speaker in unique_speakers:
        
    #     speakingTime = 0

    #     for data in json_data:
    #         if data['speaker'] == speaker:
    #             speakingTime = speakingTime + (data['end'] - data['start'])
                
    #     st.append(speakingTime)
    #     totalTime = sum(st)

    #     speakingPercentage=[]
    #     for t in st:
    #         speakingPercentage.append(round(t/totalTime*100))

    # print(st)
    # header_label1=('speaker', 'time (%)')

    # wb.create_sheet('Speaker Talktime Distribution') # Create a new sheet
    # wb.active = wb['Speaker Talktime Distribution'] # Set it as active
    # sheet1 = wb.active

    # for index, val in enumerate(header_label1):
    #     sheet1.cell(row=1, column=index+1).value = val

    # for row, speakerTime in enumerate(speakingPercentage):
    #     sheet1.cell(row+2, column=1).value = unique_speakers[row]
    #     sheet1.cell(row+2, column=2).value = speakerTime





    # Create sheet for average decibel levels per sentence.
    wb.create_sheet('Average Decibel Level') # Create a new sheet
    wb.active = wb['Average Decibel Level'] # Set it as active
    sheet2 = wb.active

    row_num1 = 2

    sheet2.cell(row=1, column=1).value = 'Sentence'
    sheet2.cell(row=1, column=2).value = 'Average dbfs'

    for data in json_data:
        avgDecibel = []
        #print(data['text'])
        sheet2.cell(row=row_num1, column=1).value = data['text']

        for dbfs in data['dbfs']:
            avgDecibel.append(dbfs['value'])

        avg = sum(avgDecibel) / len(avgDecibel)
        
        sheet2.cell(row=row_num1, column=2).value = avg
        #print(avg)

        row_num1 = row_num1 + 1


    # Set the first sheets as active.
    wb.active = wb['Speaker Diarisation']

    wb.save(name + ".xlsx")  

    return True


def unique(list):
 
    # initialize a null list
    unique_list = []
 
    # traverse for all elements
    for x in list:
        # check if exists in unique_list or not
        if x not in unique_list:
            unique_list.append(x)
   
    return unique_list